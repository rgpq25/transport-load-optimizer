import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime
import itertools
import pandas as pd
import os
import subprocess
from config import algorithm_executable_path, output_dispatches_path, output_metadata_path, input_folder_path, output_folder_path

def runAlgorithm(inputPath, algo, paramEntries):
    args = [algorithm_executable_path, "--input", inputPath, "--algo", algo]
    for k, v in paramEntries.items():
        args += [f"--{k}", str(v)]
    
    try:
        result = subprocess.run(args, capture_output=True, text=True, check=True)
        return result.stdout.strip()
    except subprocess.CalledProcessError as e:
        raise Exception("[Error de ejecución] - Error ejecutando el algoritmo:\n{e.stderr}")
    except FileNotFoundError:
        raise Exception("[Archivo no encontrado] - No se encontró el ejecutable en la ruta: {algorithm_executable_path}")


# Definición de los valores posibles para cada parámetro
if len(sys.argv) != 2:
    raise ValueError("[Error] Debe especificar el algoritmo a calibrar: 'SA-GA' o 'GRASP'.")

algo = sys.argv[1]
nIterations = 10  # Cantidad de iteraciones a realizar
inputPath = os.path.join(input_folder_path, "input_large.txt")
saga_param_grid = {
    "population": [50, 100],
    "mutation_rate": [0.2, 0.5, 0.8],
    "t_init": [100, 500, 1000],
    "t_min": [1],
    "alpha": [0.5, 0.9]
}
grasp_param_grid = {
    "iterations": [100, 300, 500, 700],
    "k_percent": [15, 30, 45, 60],
    "alphas": ["0.1,0.5,0.9"]
}

# Obtener todas las 
if algo == "SA-GA":
    keys = list(saga_param_grid.keys())
    values = list(saga_param_grid.values())
elif algo == "GRASP":
    keys = list(grasp_param_grid.keys())
    values = list(grasp_param_grid.values())
else:
    raise ValueError(f"[Error] Algoritmo desconocido: {algo}. Debe ser 'SA-GA' o 'GRASP'.")
combinations = list(itertools.product(*values))

# Construir DataFrame con ID y descripción
data = []
result = []
for i, combo in enumerate(combinations, start=1):
    currentParams = { keys[j]: combo[j] for j in range(len(keys)) }
    paramString = ", ".join([f"{k}={v}" for k, v in currentParams.items()])
    data.append((i, currentParams))

    print(f"[Configuración] - ID: {i}/{len(combinations)}, Parámetros: {paramString}")

    for nIter in range(nIterations):
        print(f"    [Ejecución] - Ejecucion {nIter + 1} del algoritmo {algo}")
        runAlgorithm(inputPath, algo, currentParams)

        metadata_df = pd.read_csv(output_metadata_path)
        metadata = metadata_df.to_dict(orient="records")[0]

        os.remove(output_dispatches_path)
        os.remove(output_metadata_path)

        result.append([
            i,
            paramString,
            nIter + 1,
            metadata.get("fitness", "N/A"),
            metadata.get("duration", "N/A"),
        ])
df = pd.DataFrame(data, columns=["ID", "Descripción de permutación"])
result_df = pd.DataFrame(result, columns=["ID", "Descripción de permutación", "Iteración", "Fitness", "Duración"])

# Crear nuevo libro de Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resultados"
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 30

# Estilos
bold_font = Font(bold=True)
highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Agrupar por permutación
grouped = result_df.groupby(["ID", "Descripción de permutación"])
current_row = 1

global_max_fitness = result_df["Fitness"].max()
stat_records = {
    "Desviacion estandar": [],
    "Maximo": [],
    "Minimo": [],
    "Promedio": []
}

for (perm_id, desc), group in grouped:
    start_row = current_row  # Para aplicar bordes luego
    fitness_values = group["Fitness"].tolist()

    # Fila 1: ID
    ws.cell(row=current_row, column=1, value="ID Permutación").font = bold_font
    ws.cell(row=current_row, column=1).fill = highlight_fill
    ws.cell(row=current_row, column=2, value=perm_id)
    current_row += 1

    # Fila 2: Descripción
    ws.cell(row=current_row, column=1, value="Descripción de permutación").font = bold_font
    ws.cell(row=current_row, column=1).fill = highlight_fill
    ws.cell(row=current_row, column=2, value=desc).alignment = openpyxl.styles.Alignment(horizontal="right")
    current_row += 1

    # Fila 4: encabezado
    ws.cell(row=current_row, column=2, value="Fitness").font = bold_font
    ws.cell(row=current_row, column=2).fill = highlight_fill
    ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal="center")
    current_row += 1

    # Filas 5-14: valores
    for i, val in enumerate(fitness_values, 1):
        ws.cell(row=current_row, column=1, value=i)
        cell = ws.cell(row=current_row, column=2, value=val)
        if val == global_max_fitness:
            cell.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        current_row += 1

    # Fila 15: vacía y fusionada
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    current_row += 1

    # Estadísticas
    stats = {
        "Desviacion estandar": pd.Series(fitness_values).std(),
        "Maximo": max(fitness_values),
        "Minimo": min(fitness_values),
        "Promedio": sum(fitness_values) / len(fitness_values)
    }

    for label, val in stats.items():
        ws.cell(row=current_row, column=1, value=label).font = bold_font
        ws.cell(row=current_row, column=1).fill = highlight_fill
        cell = ws.cell(row=current_row, column=2, value=val)
        stat_records[label].append((val, cell))
        current_row += 1

    # Aplicar bordes al rango A:start_row hasta B:current_row - 1
    for row in range(start_row, current_row):
        for col in range(1, 3):  # Columnas A y B
            ws.cell(row=row, column=col).border = thin_border

    # Espacio entre bloques
    current_row += 3

highlight_stat_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
for stat_name, entries in stat_records.items():
    if entries:
        max_val = max(entries, key=lambda x: x[0])[0]
        for val, cell in entries:
            if val == max_val:
                cell.fill = highlight_stat_fill


# Guardar a Excel
currentTime = datetime.now().strftime("%Y%m%d_%H%M%S")
outputPath = os.path.join(output_folder_path, f"calibration_{algo}_{currentTime}.xlsx")
wb.save(outputPath)
print(f"[Éxito] Archivo generado en: {outputPath}")
