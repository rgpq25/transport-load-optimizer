import sys
import os
import subprocess
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
from config import algorithm_executable_path, output_dispatches_path, output_metadata_path, input_folder_path, output_folder_path

numeric_exp_input_folder_path = os.path.join(input_folder_path, 'used_tests/numeric_experimentation')

# === CONFIGURACIÓN ===
if len(sys.argv) != 2:
    raise ValueError("[Error] Debe especificar el algoritmo: 'SA-GA' o 'GRASP'.")

algo = sys.argv[1]
paramEntries = {
    "population": 100,
    "mutation_rate": 0.2,
    "t_init": 1000,
    "t_min": 1,
    "alpha": 0.9
} if algo == "SA-GA" else {
    "iterations": 700,
    "k_percent": 30,
    "alphas": "0.1,0.5,0.9"
}
nIterations = 20

# === FUNCIONES AUXILIARES ===
def run_algorithm(input_path, algo, param_entries):
    args = [algorithm_executable_path, "--input", input_path, "--algo", algo]
    for k, v in param_entries.items():
        args += [f"--{k}", str(v)]
    try:
        result = subprocess.run(args, capture_output=True, text=True, check=True)
        return result.stdout.strip()
    except subprocess.CalledProcessError as e:
        raise Exception(f"[Error de ejecución] - {e.stderr}")
    except FileNotFoundError:
        raise Exception(f"[Archivo no encontrado] - {algorithm_executable_path}")

# === RECOGER INPUTS ===
input_files = [f for f in os.listdir(numeric_exp_input_folder_path) if f.endswith(".txt")]
input_files.sort(key=lambda x: int(x.split('input')[1].split('.')[0]))

results_by_file = {}
global_fitness_values = []

for input_file in input_files:
    input_path = os.path.join(numeric_exp_input_folder_path, input_file)
    fitness_values = []
    for i in range(nIterations):
        print(f"[{input_file}] Iteración {i+1}/{nIterations}\n")
        res = run_algorithm(input_path, algo, paramEntries)
        print(res)
        print("\n")
        metadata_df = pd.read_csv(output_metadata_path)
        fitness = metadata_df.loc[0, "fitness"]
        fitness_values.append(fitness)
        global_fitness_values.append(fitness)
        os.remove(output_dispatches_path)
        os.remove(output_metadata_path)
    results_by_file[input_file] = fitness_values

# === CREAR EXCEL ===
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resultados"
ws.column_dimensions["A"].width = 25

bold_font = Font(bold=True)
highlight_global = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
highlight_stat = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
alignment_center = Alignment(horizontal="center")

# Crear encabezados
for col, file_name in enumerate(results_by_file.keys(), start=2):
    ws.cell(row=1, column=col, value=file_name).font = bold_font
    ws.cell(row=1, column=col).fill = header_fill
    ws.cell(row=1, column=col).alignment = alignment_center

# Insertar ejecuciones
best_fitness_per_file = {
    file: max(values) for file, values in results_by_file.items()
}
for i in range(nIterations):
    ws.cell(row=i+2, column=1, value=f"Ejecucion {i+1}").font = bold_font
    ws.cell(row=i+2, column=1).fill = header_fill
    for col, file_name in enumerate(results_by_file.keys(), start=2):
        val = results_by_file[file_name][i]
        cell = ws.cell(row=i+2, column=col, value=val)
        if val == best_fitness_per_file[file_name]:
            cell.fill = highlight_global

# Calcular estadísticas
stats_names = ["Desviacion estandar", "Minimo", "Maximo", "Promedio"]
row_offset = nIterations + 2
stat_matrix = {name: [] for name in stats_names}

ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=len(results_by_file) + 1)
row_offset += 1

for stat_idx, stat in enumerate(stats_names):
    ws.cell(row=row_offset + stat_idx, column=1, value=stat).font = bold_font
    ws.cell(row=row_offset + stat_idx, column=1).fill = header_fill

for col, file_name in enumerate(results_by_file.keys(), start=2):
    series = pd.Series(results_by_file[file_name])
    stats = [
        series.std(),
        series.min(),
        series.max(),
        series.mean()
    ]
    for idx, val in enumerate(stats):
        cell = ws.cell(row=row_offset + idx, column=col, value=val)
        stat_matrix[stats_names[idx]].append((val, cell))

# Aplicar bordes
max_row = ws.max_row
max_col = ws.max_column
for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
    for cell in row:
        cell.border = thin_border

# Ajustar anchura de columnas usadas
for col in range(2, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(col)
    ws.column_dimensions[col_letter].width = 18

# Guardar archivo
output_file = os.path.join(output_folder_path, f"numExp_{algo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
wb.save(output_file)